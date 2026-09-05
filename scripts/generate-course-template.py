"""
Builds content-templates/course-content-template.xlsx from the JSON
intermediate written by generate-course-template.ts. Run the TS script
first (it reads the live catalogue); this step only formats what it wrote.

    pnpm dlx tsx scripts/generate-course-template.ts
    python3 scripts/generate-course-template.py
"""
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = ROOT / "content-templates"

with open(OUT_DIR / "course-content-template.json", encoding="utf-8") as f:
    data = json.load(f)

columns = data["columns"]
rows = data["rows"]
tracked_fields = set(data["trackedFields"])

wb = Workbook()

# --- Sheet 1: Instructions -------------------------------------------------
instructions = wb.active
instructions.title = "Instructions"
instructions.sheet_view.showGridLines = False
instructions.column_dimensions["A"].width = 100

title_font = Font(name="Arial", size=14, bold=True)
body_font = Font(name="Arial", size=11)
bold_font = Font(name="Arial", size=11, bold=True)

lines = [
    ("SRS Academy — Master Course Content Template", title_font),
    ("", body_font),
    ("How to use this workbook", bold_font),
    ("- Edit the \"Courses\" sheet. One row per real catalogue course (23 total: the 22", body_font),
    ("  courses imported from the master spreadsheet, plus Full Stack Web Development).", body_font),
    ("- Leave a cell blank if you don't have verified information yet — never type a", body_font),
    ("  placeholder like \"TBD\" or \"Coming soon\". A blank cell is the honest state.", body_font),
    ("- Cells shaded yellow are the 7 fields SRS Academy is expected to confirm over time:", body_font),
    ("  mode, level, overview, outcomes, eligibility, certification, fees. Everything else", body_font),
    ("  (course name, code, category, duration, curriculum...) is already filled from", body_font),
    ("  confirmed sources and should only be edited to correct an actual mistake.", body_font),
    ("- \"outcomes\", \"eligibility\", and \"curriculum\" can hold multiple items — separate", body_font),
    ("  them with \" | \" (a pipe with spaces on each side) in the same cell.", body_font),
    ("- \"mode\" can be more than one value (e.g. \"Online | Offline\") if a course is truly", body_font),
    ("  offered more than one way — otherwise leave it as a single value or blank.", body_font),
    ("- \"completenessState\", \"completenessPercent\" and \"missingFields\" are computed —", body_font),
    ("  don't edit them directly; they'll be regenerated the next time this file is rebuilt", body_font),
    ("  from your edits.", body_font),
    ("", body_font),
    ("Getting your edits back into the live site", bold_font),
    ("- Send the edited CSV or this workbook back, and the corresponding entries in", body_font),
    ("  src/content/catalogue/enrichment.ts get updated to match — that's the one file", body_font),
    ("  that actually changes what visitors see, on both /programs and /light/courses.", body_font),
    ("- Course name, code, category, duration and curriculum come from the source", body_font),
    ("  spreadsheet and generally shouldn't change here — flag a correction in \"notes\"", body_font),
    ("  instead of editing them directly, so it gets reviewed rather than silently applied.", body_font),
]
for i, (text, font) in enumerate(lines, start=1):
    cell = instructions.cell(row=i, column=1, value=text)
    cell.font = font
    cell.alignment = Alignment(wrap_text=False, vertical="top")

summary = data["completenessSummary"]
row_after = len(lines) + 2
instructions.cell(row=row_after, column=1, value="Current completeness (at generation time)").font = bold_font
instructions.cell(
    row=row_after + 1,
    column=1,
    value=f"Complete: {summary['complete']}   Partial: {summary['partial']}   Minimal: {summary['minimal']}",
).font = body_font

# --- Sheet 2: Courses --------------------------------------------------------
ws = wb.create_sheet("Courses")
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A2"

header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
tracked_header_fill = PatternFill(start_color="92722A", end_color="92722A", fill_type="solid")
missing_fill = PatternFill(start_color="FFF3C4", end_color="FFF3C4", fill_type="solid")
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
cell_font = Font(name="Arial", size=10)

for col_idx, name in enumerate(columns, start=1):
    c = ws.cell(row=1, column=col_idx, value=name)
    c.font = header_font
    c.fill = tracked_header_fill if name in tracked_fields else header_fill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border

tracked_col_indexes = [i for i, name in enumerate(columns, start=1) if name in tracked_fields]

for r, row in enumerate(rows, start=2):
    for c_idx, value in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c_idx, value=value if value != "" else None)
        cell.font = cell_font
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        # Highlight a blank tracked field so a gap is visible at a glance.
        if c_idx in tracked_col_indexes and (value is None or value == ""):
            cell.fill = missing_fill

widths = {
    "slug": 34, "courseCode": 12, "courseName": 34, "category": 22, "subcategory": 22,
    "courseType": 14, "duration": 12, "mode": 14, "level": 14, "overview": 40,
    "outcomes": 40, "eligibility": 40, "certification": 34, "fees": 20, "curriculum": 50,
    "featured": 10, "status": 10, "notes": 30, "completenessState": 14,
    "completenessPercent": 10, "missingFields": 34,
}
for i, name in enumerate(columns, start=1):
    ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 18)
ws.row_dimensions[1].height = 32

wb.save(OUT_DIR / "course-content-template.xlsx")
print(f"Wrote {OUT_DIR / 'course-content-template.xlsx'}")
